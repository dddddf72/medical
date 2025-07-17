import os
import shutil
import time

import openpyxl

from Config.Config import config
from Utils.Log import log
from Utils.Path import rootPath, storagePath

# 序列号
SERIAL_NUMBER = config.serial_number

# 原始手术数据文件
PATH_ORIGINAL_SURGICAL_DATA = rootPath + "/Utils/SurgicalData.xlsx"
# 保存数据后的手术数据文件
PATH_SURGICAL_DATA = storagePath + "/SurgicalData.xlsx"
# 时间格式
TIME_FORMAT = "%Y-%m-%d %H:%M:%S"

# 表格相关常量
# 表名
SURGICAL_SHEET = "surgical"
# 设备号所在列
DEVICE_ID = 1
# 开始时间所在列
START_TIME = 2
# 结束时间所在列
END_TIME = 3
# 基准值所在列
REFERENT = 4
# 警告值所在列
WARN = 5
# 备注所在列
REMARK = 6


class SurgicalData(object):
    """
    手术工具类
    保存手术数据
    """
    def __init__(self):
        log("PATH_ORIGINAL_SURGICAL_DATA", PATH_ORIGINAL_SURGICAL_DATA)
        log("PATH_SURGICAL_DATA", PATH_SURGICAL_DATA)
        # 保存数据后的手术数据文件不存在，复制一份
        if not os.path.exists(PATH_SURGICAL_DATA):
            shutil.copyfile(PATH_ORIGINAL_SURGICAL_DATA, PATH_SURGICAL_DATA)

        # 打开文件
        try:
            self.__workbook = openpyxl.load_workbook(PATH_SURGICAL_DATA)
        except BaseException as error:
            log("error SurgicalData", "文件损坏打开失败，使用原始文件替代，数据已丢失")
            log(error)
            shutil.copyfile(PATH_ORIGINAL_SURGICAL_DATA, PATH_SURGICAL_DATA)
            self.__workbook = openpyxl.load_workbook(PATH_SURGICAL_DATA)

        # 获取手术数据表
        self.__surgical_sheet = self.__workbook[SURGICAL_SHEET]
        # 手术数据表最后一行
        self.__surgical_end_row = 0
        for row in self.__surgical_sheet.iter_rows():
            self.__surgical_end_row += 1

    def __now(self):
        """
        当前时间
        :return: 字符串格式的当前时间
        """
        return time.strftime(TIME_FORMAT, time.localtime(time.time()))

    def insert_surgical(self):
        """
        插入一条新数据到手术表
        """
        # 新行
        self.__surgical_end_row += 1
        log("insert_surgical", self.__surgical_end_row)

        # 保存设备id 和 开始时间
        self.__surgical_sheet.cell(self.__surgical_end_row, DEVICE_ID).value = SERIAL_NUMBER
        self.__surgical_sheet.cell(self.__surgical_end_row, START_TIME).value = self.__now()
        # 保存
        self.__workbook.save(PATH_SURGICAL_DATA)

    def save_referent(self, value):
        """
        保存基准值
        :param value: 基准值
        """
        log("save_referent", value)
        self.__save_fluorescence(REFERENT, value)

    def save_warn(self, value):
        """
        保存警告值
        :param value: 警告值
        """
        log("save_warn", value)
        self.__save_fluorescence(WARN, value)

    def __save_fluorescence(self, column, value):
        """
        保存荧光值
        :param column: 列
        :param value: 值
        """

        # 更新结束时间
        self.__surgical_sheet.cell(self.__surgical_end_row, END_TIME).value = self.__now()

        # 荧光值
        originalValue = self.__surgical_sheet.cell(self.__surgical_end_row, column).value
        if originalValue is None:
            # 该值为空，直接写入
            self.__surgical_sheet.cell(self.__surgical_end_row, column).value = "%.4f" % value
        else:
            # 不为空，需逗号隔开
            self.__surgical_sheet.cell(self.__surgical_end_row, column).value = "%s, %.4f" % (originalValue, value)

        # 保存
        self.__workbook.save(PATH_SURGICAL_DATA)

    def add_remark(self, remark):
        """
        新增备注
        :param remark: 备注
        """
        # 原备注
        originalValue = self.__surgical_sheet.cell(self.__surgical_end_row, REMARK).value
        if originalValue is None:
            # 该值为空，直接写入
            self.__surgical_sheet.cell(self.__surgical_end_row, REMARK).value = remark
        else:
            # 不为空，需逗号隔开
            self.__surgical_sheet.cell(self.__surgical_end_row, REMARK).value = "%s, %s" % (originalValue, remark)

            # 保存
        self.__workbook.save(PATH_SURGICAL_DATA)


# 单例模式，import此属性
surgicalData = SurgicalData()
